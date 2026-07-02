"""Read a generated CTN workbook back into a :class:`ComputedDocument`.

This supports the *optional* manual path for the Tally-JSON feature: the user
downloads the generated Excel, may hand-edit it (most importantly filling the
``AS PER TALLY NAME`` column, ``D``, with the canonical Tally stock-item name),
then re-uploads it to produce the JSON. The primary path builds the JSON from
the in-memory computed result and does not use this module.

Only the columns the Tally export needs are read (names, quantities, and the
already-computed landed cost / duty / IGST amounts), reconstructing the same
``ComputedLine`` fields the exporter consumes. Reading uses openpyxl's
``data_only`` mode so cached formula results are picked up; when a workbook was
never opened/saved by Excel its formula cells have no cached value, which is
reported so the user knows to open+save (or use the direct path) instead.
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from boe_converter.excel_writer import (
    COL_AMOUNT,
    COL_AS_PER_TALLY_NAME,
    COL_CUSTOM_ASS_VALUE,
    COL_DESCRIPTION,
    COL_GST,
    COL_HSN_CODE,
    COL_LAND_COST_WITHOUT_GST,
    COL_QTY,
    COL_RATE_OF_DUTY_IGST,
    COL_RATE_PER_UNIT,
    COL_SR_NO,
    COL_TOTAL_CUSTOM_DUTY,
    COL_UNIT,
    ITEM_TABLE_FIRST_DATA_ROW,
)
from boe_converter.models import (
    ComputedDocument,
    ComputedLine,
    HeaderBlock,
    LineItem,
    RawValue,
    Totals,
)


class ExcelReadError(ValueError):
    """Raised when a workbook cannot be interpreted as a CTN sheet."""


def _rv(value) -> RawValue:
    """Wrap a cell value as a RawValue (missing when the cell is empty)."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return RawValue.missing()
    if isinstance(value, str):
        return RawValue(raw_text=value.strip(), parsed=value.strip())
    return RawValue(raw_text=str(value), parsed=value)


def _num(value) -> float | None:
    """Interpret a cell as a number, else None (blank/formula/text)."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "").strip())
        except ValueError:
            return None
    return None


def read_workbook(raw: bytes) -> ComputedDocument:
    """Reconstruct a ``ComputedDocument`` from generated CTN workbook bytes.

    Raises :class:`ExcelReadError` when no data rows carry usable numeric
    values (e.g. an unsaved formula-only workbook, or a non-CTN file).
    """
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active

    header = _read_header(ws)

    lines: list[ComputedLine] = []
    row = ITEM_TABLE_FIRST_DATA_ROW
    blank_streak = 0
    numeric_seen = False
    while row < 1000:
        sr = ws.cell(row=row, column=COL_SR_NO).value
        desc = ws.cell(row=row, column=COL_DESCRIPTION).value
        tally_name = ws.cell(row=row, column=COL_AS_PER_TALLY_NAME).value
        if not _is_serial(sr) and not (isinstance(desc, str) and desc.strip()):
            blank_streak += 1
            if blank_streak >= 3:
                break
            row += 1
            continue
        blank_streak = 0

        line, has_num = _read_line(ws, row, tally_name, desc)
        numeric_seen = numeric_seen or has_num
        lines.append(line)
        row += 1

    if not lines or not numeric_seen:
        raise ExcelReadError(
            "No usable numeric values were found in the workbook. If it was just "
            "generated, open it in Excel and save once (so formula results are "
            "cached), or use the direct 'Generate JSON' option."
        )

    totals = _totals_from_lines(lines)
    return ComputedDocument(header=header, lines=lines, totals=totals, flags=[])


def _is_serial(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and value >= 1


def _read_header(ws) -> HeaderBlock:
    def cell(ref: str) -> RawValue:
        return _rv(ws[ref].value)

    usd_rate = _num(ws["G2"].value) or 0.0
    company = ws["E1"].value
    return HeaderBlock(
        company_name=company.strip() if isinstance(company, str) else "",
        party_name=cell("E2"),
        usd_rate=usd_rate,
        details=cell("E3"),
        invoice_no=cell("E4"),
        invoice_date=cell("G4"),
        be_no=cell("E5"),
        be_date=cell("G5"),
        bl_no=cell("E6"),
        bl_date=cell("G6"),
        invoice_amount=RawValue.missing(),
        invoice_currency=RawValue.missing(),
        package_count=RawValue.missing(),
        container_details=RawValue.missing(),
    )


def _read_line(ws, row: int, tally_name, desc) -> tuple[ComputedLine, bool]:
    """Build a ComputedLine from one Excel row; also report if it had numbers.

    The stock-item name prefers the hand-filled ``AS PER TALLY NAME`` (column D)
    and falls back to the ``Description`` (column E).
    """
    def v(col: int):
        return ws.cell(row=row, column=col).value

    name = tally_name if (isinstance(tally_name, str) and tally_name.strip()) else desc
    serial = v(COL_SR_NO)
    try:
        item_serial = int(serial) if _is_serial(serial) else row
    except (TypeError, ValueError):
        item_serial = row

    igst_rate = _num(v(COL_RATE_OF_DUTY_IGST))
    source = LineItem(
        item_serial=item_serial,
        cth_hsn=_rv(v(COL_HSN_CODE)),
        description=_rv(name),
        unit_price_usd=RawValue.missing(),
        quantity=_rv(v(COL_QTY)),
        unit=_rv(v(COL_UNIT)),
        assessable_value=_rv(v(COL_CUSTOM_ASS_VALUE)),
        bcd_rate=RawValue.missing(),
        bcd_amount=RawValue.missing(),
        igst_rate=RawValue(raw_text=str(igst_rate), parsed=igst_rate)
        if igst_rate is not None
        else RawValue.missing(),
        total_duty=RawValue.missing(),
    )

    # Column L "Amount" is the USD amount, not the INR purchase value. The INR
    # purchase (party) value is land-cost-excl-GST minus the customs duty
    # (land_cost_excl_gst = purchase_inr + total_customs_duty).
    land_excl = _num(v(COL_LAND_COST_WITHOUT_GST))
    duty = _num(v(COL_TOTAL_CUSTOM_DUTY))
    igst_amt = _num(v(COL_GST))
    rate_per_unit = _num(v(COL_RATE_PER_UNIT))
    amount_usd = _num(v(COL_AMOUNT))
    purchase_inr = (
        land_excl - duty if land_excl is not None and duty is not None else None
    )
    has_num = any(x is not None for x in (purchase_inr, land_excl, duty, igst_amt))

    line = ComputedLine(
        source=source,
        purchase_inr=purchase_inr,
        total_customs_duty=duty,
        igst_amount=igst_amt,
        land_cost_excl_gst=land_excl,
        purchase_rate_per_unit=rate_per_unit,
    )
    return line, has_num


def _totals_from_lines(lines: list[ComputedLine]) -> Totals:
    return Totals(
        total_assessable_value=sum(_num(l.source.assessable_value) or 0.0 for l in lines),
        total_customs_duty=sum(l.total_customs_duty or 0.0 for l in lines),
        total_igst=sum(l.igst_amount or 0.0 for l in lines),
        total_land_cost_excl_gst=sum(l.land_cost_excl_gst or 0.0 for l in lines),
    )


def write_tally_names(raw: bytes, names: dict[int, str]) -> bytes:
    """Fill the ``AS PER TALLY NAME`` column (D) of a generated CTN workbook.

    ``names`` maps a line's serial number (column ``A``) to the canonical Tally
    stock-item name to write in column ``D`` of that row. Rows whose serial is
    not in ``names`` are left untouched. Returns the edited workbook bytes.

    This lets Step 2 (name mapping) round-trip the *same* perfectly-formatted
    workbook produced by Step 1 - only column D is edited, nothing else changes.
    """
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    row = ITEM_TABLE_FIRST_DATA_ROW
    blank_streak = 0
    while row < 5000:
        sr = ws.cell(row=row, column=COL_SR_NO).value
        desc = ws.cell(row=row, column=COL_DESCRIPTION).value
        if not _is_serial(sr) and not (isinstance(desc, str) and desc.strip()):
            blank_streak += 1
            if blank_streak >= 3:
                break
            row += 1
            continue
        blank_streak = 0
        if _is_serial(sr):
            try:
                serial = int(sr)
            except (TypeError, ValueError):
                serial = None
            if serial is not None and serial in names and names[serial].strip():
                ws.cell(row=row, column=COL_AS_PER_TALLY_NAME).value = names[serial].strip()
        row += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
