"""Excel_Generator: writes the ComputedDocument into the exact CTN layout.

Milestone 1. Builds ``Sheet1`` in the exact ``1357 ctn llp.xlsx`` layout:

- Header_Block in ``D1:G8`` (labels in column D/F, values in column E/G).
- Item_Table header at row 12 (``A12:Y12``), with each label reproduced
  character-for-character from the sample (including its quirks, e.g.
  ``'Rate Per USDin purcahse '`` and ``'Ratepurchase per unitPcs/KGS/SET'``).
- Line items written from row 13, one dense row per item in ascending serial
  order with no blank rows, ``Sr. no.`` running 1..N.

Computed values are written as full-precision literals (Req 8.5); directly
extracted values are written verbatim (Req 8.4). Review-flagged / missing cells
and no-source columns are left blank (Req 5.11, 6.13, 4.8, 4.9, 8.6); a field
that was located but could not be parsed is written as its raw text (Req 9.2).

The Totals_Row (row 61) and the auxiliary template sections are written by
``_write_totals_row`` / ``_write_aux_templates`` (task 7.2). The totals row
reproduces the sample's summed columns (``G, L, M, N, O, P, Q, S, U, W, X``)
plus the BOE package count in column G; the auxiliary sections reproduce the
``DETAILS AS PER CHALLANS`` / ``DETAILS AS PER TALLY`` / ``CLEARANCE AND
FORWARDING INVOICE`` and C&F block labels verbatim with empty data cells.

Cell map grounded directly in the sample workbook (design.md "Target Excel cell
map"):

    A Sr. no.      (computed 1..N)        N CUSTOM ASS VALUE            (direct)
    B PARTY NAME   (blank)                O LAND COST ... WITHOUT GST   (computed)
    C BILLING AMOUNT (blank)              P TOTAL Custom Duty           (computed)
    D AS PER TALLY NAME (blank)           Q GST                         (computed)
    E Description  (direct)               R RATE OF DUTY IGST           (direct rate)
    F HSN CODE     (direct)               S total custom duty           (computed)
    G CTN          (blank)                T RATE OF INTEREST            (direct BCD rate)
    H QTY          (direct)               U CUST AIDC                   (direct BCD amount)
    I Unit         (direct)               V RATE OF INTEREST            (constant 0.10 SWS)
    J pcs          (computed)             W SURCHARGE                   (computed SWS)
    K Unit Price in USD (direct)          X LAND COST ... WITH GST      (computed)
    L Amount       (computed)             Y Ratepurchase per unit...    (computed)
    M Rate Per USDin purcahse  (computed)
"""

from __future__ import annotations

from copy import copy
from functools import lru_cache
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from boe_converter.models import (
    ComputedDocument,
    ComputedLine,
    HeaderBlock,
    RawValue,
    ReviewFlagSet,
    Totals,
)

# The bundled golden workbook used as the *style template*: it carries the
# sample's exact column widths, row heights, merged ranges, fonts, fills,
# borders, number formats and the static decorative band / auxiliary
# scaffolding. The generator loads it as the base workbook, clears the
# document-specific cells, and writes this document's values into the
# already-styled cells, so the output is visually identical to
# ``1357 ctn llp.xlsx`` (Req 8.x). Only data is replaced; styling is inherited.
_TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "ctn_template.xlsx"

# Template layout anchors (the sample's fixed structure).
_TEMPLATE_FIRST_DATA_ROW = 13
_TEMPLATE_LAST_DATA_ROW = 60          # styled data region rows 13..60
_TEMPLATE_TOTALS_ROW = 61
# A representative interior data row whose per-column styling (borders, number
# formats, fonts) is cloned onto continuation rows when a document has more
# line items than the template's styled data region (overflow handling).
_TEMPLATE_STYLE_ROW = 14


@lru_cache(maxsize=1)
def _template_bytes() -> bytes:
    """Read and cache the template workbook bytes (avoids repeated disk I/O)."""
    return _TEMPLATE_PATH.read_bytes()


def _load_template() -> Workbook:
    """Return a fresh, writable copy of the style template workbook."""
    return load_workbook(BytesIO(_template_bytes()))

# --- Item_Table column indices (1-based, A=1 .. Y=25) ----------------------
COL_SR_NO = 1                       # A  Sr. no.            (computed sequential)
COL_PARTY_NAME = 2                  # B  PARTY NAME         (blank, no source)
COL_BILLING_AMOUNT = 3              # C  BILLING AMOUNT     (blank, no source)
COL_AS_PER_TALLY_NAME = 4           # D  AS PER TALLY NAME  (blank, Manual/External)
COL_DESCRIPTION = 5                 # E  Description        (direct)
COL_HSN_CODE = 6                    # F  HSN CODE           (direct)
COL_CTN = 7                         # G  CTN                (blank, Manual/External)
COL_QTY = 8                         # H  QTY                (direct)
COL_UNIT = 9                        # I  Unit               (direct)
COL_PCS = 10                        # J  pcs                (computed)
COL_UNIT_PRICE_USD = 11            # K  Unit Price in USD   (direct)
COL_AMOUNT = 12                     # L  Amount             (computed)
COL_RATE_PER_USD = 13              # M  Rate Per USDin purcahse  (computed)
COL_CUSTOM_ASS_VALUE = 14         # N  CUSTOM ASS VALUE    (direct)
COL_LAND_COST_WITHOUT_GST = 15    # O  LAND COST ... WITHOUT GST (computed)
COL_TOTAL_CUSTOM_DUTY = 16        # P  TOTAL Custom Duty   (computed)
COL_GST = 17                       # Q  GST                (computed)
COL_RATE_OF_DUTY_IGST = 18        # R  RATE OF DUTY IGST   (direct igst rate)
COL_TOTAL_CUSTOM_DUTY_2 = 19      # S  total custom duty   (computed combined duty)
COL_RATE_OF_INTEREST_BCD = 20     # T  RATE OF INTEREST    (direct BCD rate)
COL_CUST_AIDC = 21                 # U  CUST AIDC          (direct BCD amount)
COL_RATE_OF_INTEREST_SWS = 22     # V  RATE OF INTEREST    (constant 0.10 SWS rate)
COL_SURCHARGE = 23                 # W  SURCHARGE          (computed SWS amount)
COL_LAND_COST_WITH_GST = 24       # X  LAND COST ... WITH GST (computed)
COL_RATE_PER_UNIT = 25            # Y  Ratepurchase per unitPcs/KGS/SET (computed)

ITEM_TABLE_HEADER_ROW = 12
ITEM_TABLE_FIRST_DATA_ROW = 13

# The Totals_Row's default position (row 61), used when the document fits within
# the template's styled data region (<=48 items). When a document has more line
# items than the region, the totals row and the auxiliary sections below it are
# shifted down by the overflow amount so they never collide with item rows
# (dynamic positioning); see ``_resolve_layout`` / ``_expand_for_overflow``.
TOTALS_ROW = 61

# The constant SWS rate written into column V for every line item (sample V=0.1).
SWS_RATE = 0.10

# ---------------------------------------------------------------------------
# Auxiliary template labels, captured character-for-character from the sample
# workbook (``1357 ctn llp.xlsx``). Keyed by cell coordinate -> verbatim label.
# Trailing spaces, double spaces and quirks (e.g. 'EWAY BILL N0' with a zero,
# 'BL  NO. ' with two spaces) are deliberate and must be preserved (Req 8.3).
# Only the labels are written; the data cells within these sections are left
# empty in Milestone 1 (data entry / Tally linkage deferred to Milestone B).
# ---------------------------------------------------------------------------
AUX_LABELS: dict[str, str] = {
    # 'total usd' summary label (its value cell E71 is written separately as the
    # computed total USD, mirroring the sample's '=L61').
    "B71": "total usd",
    # DETAILS AS PER CHALLANS section title + column header row (row 72).
    "G71": "DETAILS AS PER CHALLANS",
    "G72": "MARKA",
    "H72": "PARTY NAME",
    "I72": "CTN",
    "J72": "DATE ",
    "K72": "CHALLAN NO.",
    "L72": "TAXABLE",
    "M72": "GST VALUE",
    "N72": "TOTAL CHALLAN AMT.",
    "O72": "GST RATE",
    # DETAILS AS PER TALLY section title + column header row (row 72).
    "Q71": "DETAILS AS PER TALLY",
    "Q72": "SR NO",
    "R72": "PARTY NAME",
    "S72": "DATE ",
    "T72": "BILL NO",
    "U72": "TAXABLE VALUE",
    "V72": "GST VALUE",
    "W72": "ROUND OFF",
    "X72": "TOTAL VALUE",
    # C&F detail block - row labels in column B (B72-B86); data cells (E*) empty.
    "B72": "customer and other expenses",
    "B73": "IGST",
    "B74": "C&F AGENCY",
    "B75": "B/E NO",
    "B76": "B/E DATE",
    "B77": "BL  NO. ",
    "B78": "BL DATE",
    "B79": "NET WEIGHT",
    "B80": "GROSS WEIGHT",
    "B81": "agst bill no ",
    "B82": "INVOICE RATE  IN USD",
    "B83": "EWAY BILL N0",
    "B84": "CHA CO NAME",
    "B85": "remdince rate",
    "B86": "remdince date",
    # CLEARANCE AND FORWARDING INVOICE section title + column header row (row 95).
    "G94": "CLEARANCE AND FORWARDING INVOICE",
    "G95": "SR NO",
    "H95": "PARTY NAME",
    "I95": "DATE ",
    "J95": "INVOICE NO",
    "K95": "NET AMOUNT",
    "L95": "GST VALUE",
    "M95": "TDS",
    "N95": "ROUND OFF",
    "O95": "TOTALAMOUNT",
}

# The 'total usd' value cell (B71 label -> E71 value), mirroring sample '=L61'.
AUX_TOTAL_USD_CELL = "E71"

# Item_Table header labels, reproduced character-for-character from the sample
# (Req 5.12). Keyed by 1-based column index. Spacing/typo quirks are deliberate.
ITEM_TABLE_HEADERS: dict[int, str] = {
    COL_SR_NO: "Sr. no.",
    COL_PARTY_NAME: "PARTY NAME",
    COL_BILLING_AMOUNT: "BILLING AMOUNT",
    COL_AS_PER_TALLY_NAME: "AS PER TALLY NAME",
    COL_DESCRIPTION: "Description",
    COL_HSN_CODE: "HSN CODE",
    COL_CTN: "CTN",
    COL_QTY: "QTY",
    COL_UNIT: "Unit",
    COL_PCS: "pcs",
    COL_UNIT_PRICE_USD: "Unit Price in USD",
    COL_AMOUNT: "Amount",
    COL_RATE_PER_USD: "Rate Per USDin purcahse ",
    COL_CUSTOM_ASS_VALUE: "CUSTOM ASS VALUE",
    COL_LAND_COST_WITHOUT_GST: "LAND COST OF PURCHASE WITHOUT GST",
    COL_TOTAL_CUSTOM_DUTY: "TOTAL Custom Duty",
    COL_GST: "GST",
    COL_RATE_OF_DUTY_IGST: "RATE OF DUTY IGST",
    COL_TOTAL_CUSTOM_DUTY_2: "total custom duty",
    COL_RATE_OF_INTEREST_BCD: "RATE OF INTEREST",
    COL_CUST_AIDC: "CUST AIDC",
    COL_RATE_OF_INTEREST_SWS: "RATE OF INTEREST",
    COL_SURCHARGE: "SURCHARGE",
    COL_LAND_COST_WITH_GST: "LAND COST OF PURCHASE WITH GST",
    COL_RATE_PER_UNIT: "Ratepurchase per unitPcs/KGS/SET",
}


def _raw_cell_value(rv: RawValue | None) -> object | None:
    """Resolve the value to write for a directly-extracted ``RawValue`` cell.

    - ``None`` / missing -> ``None`` (the cell is left blank; Req 4.9, 5.11).
    - Unparseable (located but not resolvable) -> the raw text (Req 9.2): the
      printed characters are preserved rather than dropped.
    - Otherwise the parsed interpretation is written when present (full numeric
      precision / verbatim string; Req 8.4), falling back to the raw text.
    """
    if rv is None or rv.is_missing:
        return None
    if rv.is_unparseable:
        return rv.raw_text
    if rv.parsed is not None:
        return rv.parsed
    return rv.raw_text


def _raw_number(rv: RawValue | None) -> float | None:
    """Return the numeric interpretation of a ``RawValue`` (or ``None``).

    Used to total per-line columns whose sums are not pre-computed on ``Totals``
    (e.g. column U ``CUST AIDC`` = BCD amount). Mirrors the column's own written
    value: missing/unparseable/non-numeric yields ``None`` (contributes nothing
    to the sum), preserving full precision otherwise.
    """
    if rv is None or rv.is_missing or rv.is_unparseable:
        return None
    parsed = rv.parsed
    if isinstance(parsed, bool):
        return None
    if isinstance(parsed, (int, float)):
        return float(parsed)
    return None


def _sum_optional(values) -> float:
    """Sum ``float | None`` values, skipping ``None`` (blank cells).

    ``None`` per-line values (a required input was missing/non-numeric, Req 6.13)
    do not contribute. An empty/all-``None`` iterable yields ``0.0`` (Req 7.3).
    Kept at full floating-point precision, matching the sample's ``SUM`` cells.
    """
    total = 0.0
    for value in values:
        if value is not None:
            total += value
    return total


def _shift_coord(coordinate: str, shift: int) -> str:
    """Return ``coordinate`` with its row shifted down by ``shift`` rows.

    Used to relocate the auxiliary-section labels (and the ``total usd`` value
    cell) when the line-item count overflows the template's styled data region.
    A non-positive ``shift`` returns the coordinate unchanged, so documents that
    fit the template keep the sample's exact cell positions (Req 8.2/8.3).
    """
    if shift <= 0:
        return coordinate
    row, col = coordinate_to_tuple(coordinate)
    return f"{get_column_letter(col)}{row + shift}"


class ExcelGenerator:
    """Builds the CTN-layout workbook and returns ``.xlsx`` bytes."""

    def __init__(self, use_formulas: bool = True) -> None:
        """Create a generator.

        ``use_formulas`` (default ``True``) makes the computed cells carry live
        Excel formulas mirroring the sample workbook (e.g. ``L=H*K``,
        ``M=L*<usd_rate>``, ``=SUM(...)`` totals) rather than pre-evaluated
        literal values, so the downloaded workbook recalculates in Excel exactly
        like ``1357 ctn llp.xlsx``. Set it to ``False`` to emit full-precision
        literal values instead (used by tests that read the evaluated numbers
        back without an Excel engine to recalculate them).
        """
        self.use_formulas = use_formulas

    def generate(self, doc: ComputedDocument, flags: ReviewFlagSet) -> bytes:
        """Build ``Sheet1`` in the exact CTN layout and return ``.xlsx`` bytes.

        Constructs the workbook in memory (header block + item table) and
        serialises it. The Totals_Row and auxiliary templates are added by
        task 7.2's ``_write_totals_row`` / ``_write_aux_templates``.
        """
        wb = self.build_workbook(doc, flags)
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def build_workbook(self, doc: ComputedDocument, flags: ReviewFlagSet) -> Workbook:
        """Assemble the workbook from the bundled CTN style template.

        Loads ``ctn_template.xlsx`` (which carries the sample's exact column
        widths, row heights, merges, fonts, fills, borders and number formats),
        clears its dynamic cells, then writes this document's header values,
        line items and totals into the already-styled cells. The result is
        visually identical to ``1357 ctn llp.xlsx`` while carrying this
        document's data. Exactly one sheet named ``Sheet1`` (Req 8.1).

        When the document has more line items than the template's styled data
        region (rows 13..60), the totals row and the auxiliary sections below it
        are shifted down so they never overlap the item rows; the additional
        item rows inherit the data-row styling (see ``_expand_for_overflow``).
        """
        wb = _load_template()
        ws = wb["Sheet1"]

        n_items = len(doc.lines)
        last_data_row = _TEMPLATE_FIRST_DATA_ROW + n_items - 1
        # Overflow: how far the data extends past the template's styled region.
        shift = max(0, last_data_row - _TEMPLATE_LAST_DATA_ROW)
        if shift:
            self._expand_for_overflow(ws, shift)
        totals_row = _TEMPLATE_TOTALS_ROW + shift

        self._prepare_template(ws, n_items, totals_row)
        self._write_header_block(ws, doc.header, flags)
        self._write_item_table(ws, doc.lines, doc.header.usd_rate, flags)
        self._write_totals_row(ws, doc.totals, doc.lines, totals_row)
        self._write_aux_templates(ws, shift, totals_row)
        return wb

    # ------------------------------------------------------------------
    # Overflow handling (shift totals + aux sections below the item rows)
    # ------------------------------------------------------------------
    def _expand_for_overflow(self, ws: Worksheet, shift: int) -> None:
        """Make room for ``shift`` extra line-item rows below the data region.

        Inserts ``shift`` rows immediately before the template totals row, which
        pushes the totals row and every auxiliary section below it down by
        ``shift`` (their values and styles move with them). openpyxl does not
        relocate merged ranges on insert, so the aux merges (``G71:O71``,
        ``Q71:X71``, ``G94:O94``) are re-created at their shifted positions. The
        freshly inserted rows arrive unstyled, so each one is given the data-row
        styling cloned from a representative template data row, keeping the new
        item rows visually identical to the originals.
        """
        # Capture and release the merges that sit at/below the totals row before
        # inserting. They must be unmerged while their cells still exist (openpyxl
        # does not relocate merge coordinates on insert, so unmerging by the old
        # coordinates afterwards would fail once the cells have shifted).
        below_specs = [
            (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
            for rng in list(ws.merged_cells.ranges)
            if rng.min_row >= _TEMPLATE_TOTALS_ROW
        ]
        for min_row, min_col, max_row, max_col in below_specs:
            ws.unmerge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col,
            )

        ws.insert_rows(_TEMPLATE_TOTALS_ROW, shift)

        # Re-create the captured merges at their shifted positions.
        for min_row, min_col, max_row, max_col in below_specs:
            ws.merge_cells(
                start_row=min_row + shift,
                start_column=min_col,
                end_row=max_row + shift,
                end_column=max_col,
            )

        # Style the newly inserted continuation rows like a real data row.
        self._clone_data_row_style(
            ws, _TEMPLATE_TOTALS_ROW, _TEMPLATE_TOTALS_ROW + shift - 1
        )

    @staticmethod
    def _clone_data_row_style(ws: Worksheet, first_row: int, last_row: int) -> None:
        """Copy the data-row styling onto rows ``first_row..last_row`` (inclusive).

        Clones the per-column style (borders, fonts, fills, number formats) and
        the row height from ``_TEMPLATE_STYLE_ROW`` so overflow item rows render
        identically to the template's styled data region. Values are left blank;
        the item writer fills them in afterwards.
        """
        ref_height = ws.row_dimensions[_TEMPLATE_STYLE_ROW].height
        for row in range(first_row, last_row + 1):
            if ref_height is not None:
                ws.row_dimensions[row].height = ref_height
            for col in range(1, COL_RATE_PER_UNIT + 1):
                src = ws.cell(row=_TEMPLATE_STYLE_ROW, column=col)
                dst = ws.cell(row=row, column=col)
                dst._style = copy(src._style)
                dst.value = None

    # ------------------------------------------------------------------
    # Template preparation (clear the sample's data, keep its styling)
    # ------------------------------------------------------------------
    def _prepare_template(self, ws: Worksheet, n_items: int, totals_row: int) -> None:
        """Clear the template's document-specific cells, preserving all styling.

        Removes the sample's baked-in header values, line items and totals so
        only this document's data is written, while leaving every style, width,
        height and the auxiliary scaffolding intact. Merged ranges that fall in
        the line-item write region are unmerged so each item keeps its own row.
        ``totals_row`` is the (possibly shifted) row the totals are written to;
        the data region and that row are cleared together.
        """
        last_data_row = _TEMPLATE_FIRST_DATA_ROW + max(n_items, 0) - 1
        write_end = max(_TEMPLATE_LAST_DATA_ROW, last_data_row)

        # Unmerge any merged range intersecting the data write region (e.g. the
        # sample's G36:G57) so per-row writes never target a merged cell.
        for rng in list(ws.merged_cells.ranges):
            if not (rng.max_row < _TEMPLATE_FIRST_DATA_ROW or rng.min_row > write_end):
                ws.unmerge_cells(str(rng))

        # Clear the header block value cells (labels are rewritten by the header
        # writer); no-source value cells simply stay blank.
        for coord in (
            "E1", "E2", "G2", "E3", "G3", "E4", "G4",
            "E5", "G5", "E6", "G6", "E7", "G7", "E8", "G8",
        ):
            ws[coord] = None

        # Clear the data region and the totals row (cols A..Y), keeping styles.
        for row in range(_TEMPLATE_FIRST_DATA_ROW, max(write_end, totals_row) + 1):
            for col in range(1, COL_RATE_PER_UNIT + 1):
                ws.cell(row=row, column=col).value = None

        # Harmonize the styled data region. The bundled template carries a few
        # hand-edited rows (58-60) with stray bold / missing-highlight quirks
        # left over from the original sample; cloning the canonical data row's
        # per-column style across the whole region makes every item row render
        # identically (Req 8.x visual fidelity).
        self._normalize_data_region_style(ws)

    @staticmethod
    def _normalize_data_region_style(ws: Worksheet) -> None:
        """Clone the canonical data-row style across rows 13..60.

        Copies the per-column style (fonts, fills, borders, number formats) from
        ``_TEMPLATE_STYLE_ROW`` onto every row of the template's styled data
        region so the hand-edited inconsistencies in the sample's last rows
        (e.g. row 60's stray bold, rows 58-59's missing column-O highlight) do
        not surface when those rows carry item data. Values are untouched.
        """
        for row in range(_TEMPLATE_FIRST_DATA_ROW, _TEMPLATE_LAST_DATA_ROW + 1):
            if row == _TEMPLATE_STYLE_ROW:
                continue
            for col in range(1, COL_RATE_PER_UNIT + 1):
                src = ws.cell(row=_TEMPLATE_STYLE_ROW, column=col)
                ws.cell(row=row, column=col)._style = copy(src._style)

    # ------------------------------------------------------------------
    # Header block (D1:G8)
    # ------------------------------------------------------------------
    def _write_header_block(
        self, ws: Worksheet, header: HeaderBlock, flags: ReviewFlagSet
    ) -> None:
        """Write the Header_Block labels and values into ``D1:G8`` (Req 4.x).

        Labels (column D/F) are reproduced verbatim from the sample. Values
        (column E/G) are written exactly as extracted, without reformatting the
        number/date values (Req 4.2-4.5). A value flagged missing/unreadable, and
        every no-source field (``Eway bill no/date``, ``RETTENCE DATE/RATE``,
        ``USD Amt``), is left blank with no characters or placeholder (Req 4.8,
        4.9, 8.6).
        """
        # --- fixed labels (column D and F) ---
        ws["D1"] = "Company name"
        ws["D2"] = "Party Name"
        ws["F2"] = "USD Rate"
        ws["D3"] = "Details"
        ws["F3"] = "USD Amt"          # no-source label; value cell stays blank
        ws["D4"] = "Invoice No"
        ws["F4"] = "Inv Date"
        ws["D5"] = "BE No"
        ws["F5"] = "BE Date"
        ws["D6"] = "B/L NO"
        ws["F6"] = "B/L DATE"
        ws["D7"] = "Eway bill no"     # no-source label; value cell stays blank
        ws["F7"] = "Eway bill date"   # no-source label; value cell stays blank
        ws["D8"] = "RETTENCE DATE"    # no-source label; value cell stays blank
        ws["F8"] = "RETTENCE RATE"    # no-source label; value cell stays blank

        # --- values (column E and G) ---
        # Company name is a configuration string (Req 4.7), always present.
        if header.company_name:
            ws["E1"] = header.company_name

        self._set_if_value(ws, "E2", _raw_cell_value(header.party_name))   # Req 4.1
        # USD Rate is User-supplied; write at its full numeric precision (Req 4.5).
        if header.usd_rate is not None:
            ws["G2"] = header.usd_rate
        self._set_if_value(ws, "E3", _raw_cell_value(header.details))      # Req 4.6
        self._set_if_value(ws, "E4", _raw_cell_value(header.invoice_no))   # Req 4.2
        self._set_if_value(ws, "G4", _raw_cell_value(header.invoice_date)) # Req 4.2
        self._set_if_value(ws, "E5", _raw_cell_value(header.be_no))        # Req 4.3
        self._set_if_value(ws, "G5", _raw_cell_value(header.be_date))      # Req 4.3
        self._set_if_value(ws, "E6", _raw_cell_value(header.bl_no))        # Req 4.4
        self._set_if_value(ws, "G6", _raw_cell_value(header.bl_date))      # Req 4.4

    # ------------------------------------------------------------------
    # Item table (header row 12, data from row 13)
    # ------------------------------------------------------------------
    def _write_item_table(
        self, ws: Worksheet, lines: list[ComputedLine], usd_rate: float, flags: ReviewFlagSet
    ) -> None:
        """Write the Item_Table header (row 12) and the dense data rows (Req 5.x).

        The header labels are reproduced character-for-character (Req 5.12). Line
        items are written in ascending item-serial order beginning at row 13, one
        per row with no blank rows (Req 5.1); ``Sr. no.`` runs 1..N (Req 5.2).
        Direct values are written verbatim and computed values at full precision
        into their mapped columns (Req 5.3-5.10, 6.11, 8.4, 8.5). Missing/
        flagged cells and no-source columns stay blank (Req 5.11, 8.6); located
        but unparseable fields are written as raw text (Req 9.2).
        """
        # 'In USD' sub-header sits above the 'Amount' column (sample L11).
        ws.cell(row=ITEM_TABLE_HEADER_ROW - 1, column=COL_AMOUNT, value="In USD")

        # Header row 12 - verbatim labels.
        for col, label in ITEM_TABLE_HEADERS.items():
            ws.cell(row=ITEM_TABLE_HEADER_ROW, column=col, value=label)

        # Data rows, ascending serial, dense (no gaps), Sr. no. 1..N.
        ordered = sorted(lines, key=lambda ln: ln.source.item_serial)
        for offset, line in enumerate(ordered):
            row = ITEM_TABLE_FIRST_DATA_ROW + offset
            self._write_item_row(ws, row, offset + 1, line, usd_rate, flags)

    def _write_item_row(
        self,
        ws: Worksheet,
        row: int,
        sr_no: int,
        line: ComputedLine,
        usd_rate: float,
        flags: ReviewFlagSet,
    ) -> None:
        """Write one Line_Item's cells into ``row`` per the column map.

        Columns B (PARTY NAME), C (BILLING AMOUNT), D (AS PER TALLY NAME) and
        G (CTN) have no BOE/configuration source in Milestone 1 and are left
        blank for every row (Req 8.6, mapping table). Review-flagged computed
        values are already ``None`` on the ``ComputedLine`` and so write blank.
        """
        item = line.source

        # A: Sr. no. - dense sequential 1..N (Req 5.2).
        ws.cell(row=row, column=COL_SR_NO, value=sr_no)

        # Direct (verbatim) values.
        self._set_cell(ws, row, COL_DESCRIPTION, _raw_cell_value(item.description))
        self._set_cell(ws, row, COL_HSN_CODE, _raw_cell_value(item.cth_hsn))
        # G: per-line CTN - populated from the supplier invoice when supplied,
        # otherwise blank (the BOE itself has no per-line carton count).
        self._set_cell(ws, row, COL_CTN, _raw_cell_value(item.cartons))
        self._set_cell(ws, row, COL_QTY, _raw_cell_value(item.quantity))
        self._set_cell(ws, row, COL_UNIT, _raw_cell_value(item.unit))
        self._set_cell(ws, row, COL_UNIT_PRICE_USD, _raw_cell_value(item.unit_price_usd))
        self._set_cell(ws, row, COL_CUSTOM_ASS_VALUE, _raw_cell_value(item.assessable_value))
        self._set_cell(ws, row, COL_RATE_OF_DUTY_IGST, _raw_cell_value(item.igst_rate))
        self._set_cell(ws, row, COL_RATE_OF_INTEREST_BCD, _raw_cell_value(item.bcd_rate))
        # U: CUST AIDC = BCD amount + CHCESS (computed); falls back to the raw
        # BCD amount when no cess is present (cust_aidc == bcd_amount).
        self._set_cell(ws, row, COL_CUST_AIDC, line.cust_aidc)

        # Constant SWS rate (column V) - sample writes 0.1 on every line.
        ws.cell(row=row, column=COL_RATE_OF_INTEREST_SWS, value=SWS_RATE)

        # GST/IGST amount (column Q) is a literal in the sample (manually keyed),
        # so it is always written as a value and is referenced by the S and X
        # formulas below.
        self._set_cell(ws, row, COL_GST, line.igst_amount)

        # Computed (full-precision) values. When formula mode is on these carry
        # the sample's live Excel formulas (referencing the literal cells H, K,
        # U, V, Q on the same row); otherwise the pre-evaluated literal is
        # written. A ``None`` computed value (missing/non-numeric input, Req
        # 6.13) leaves the cell blank in both modes so the formula never turns a
        # missing input into a spurious 0.
        qty = _raw_number(item.quantity)
        self._put_computed(ws, row, COL_PCS, line.pcs, f"=H{row}*12")
        self._put_computed(ws, row, COL_AMOUNT, line.amount_usd, f"=H{row}*K{row}")
        self._put_computed(
            ws, row, COL_RATE_PER_USD, line.purchase_inr, f"=L{row}*{self._num(usd_rate)}"
        )
        self._put_computed(ws, row, COL_SURCHARGE, line.sws_amount, f"=U{row}*V{row}")
        self._put_computed(
            ws, row, COL_TOTAL_CUSTOM_DUTY, line.total_customs_duty, f"=U{row}+W{row}"
        )
        self._put_computed(
            ws, row, COL_LAND_COST_WITHOUT_GST, line.land_cost_excl_gst, f"=M{row}+P{row}"
        )
        self._put_computed(
            ws, row, COL_TOTAL_CUSTOM_DUTY_2, line.combined_duty, f"=P{row}+Q{row}"
        )
        self._put_computed(
            ws, row, COL_LAND_COST_WITH_GST, line.land_cost_incl_gst, f"=M{row}+S{row}"
        )
        # Y = O/H; guard the qty==0 case (calculator yields 0) so the formula
        # never produces a #DIV/0! error in the workbook.
        if line.purchase_rate_per_unit is not None and self.use_formulas and qty not in (None, 0):
            ws.cell(row=row, column=COL_RATE_PER_UNIT, value=f"=O{row}/H{row}")
        else:
            self._set_cell(ws, row, COL_RATE_PER_UNIT, line.purchase_rate_per_unit)

    # ------------------------------------------------------------------
    # Totals row (row 61)
    # ------------------------------------------------------------------
    def _write_totals_row(
        self, ws: Worksheet, totals: Totals, lines: list[ComputedLine], totals_row: int
    ) -> None:
        """Write the Totals_Row at ``totals_row`` (row 61, or shifted on overflow).

        Reproduces the sample's summed columns (``G, L, M, N, O, P, Q, S, U, W,
        X``) at full floating-point precision (Req 7.4, 8.2). Columns backed by
        a pre-computed ``Totals`` field use it directly (L, N, O, P, Q, X); the
        remaining column sums (M, S, U, W) are summed here from the per-line
        values exactly as written into those columns, matching the sample's
        ``SUM`` formulas. Column G carries the BOE total package/CTN count taken
        from ``Totals.package_count`` (Req 7.5); if it is missing the cell is
        left blank rather than substituting a value.
        """
        row = totals_row
        n = len(lines)
        first = ITEM_TABLE_FIRST_DATA_ROW
        last = first + n - 1  # inclusive last data row; < first when no items

        # G: CTN total = BOE total package count (Req 7.5). Per-line CTN cells are
        # blank in Milestone 1; the total is the extracted package count and is
        # always written as a literal (never a SUM of the empty per-line column).
        self._set_cell(ws, row, COL_CTN, _raw_cell_value(totals.package_count))

        # The remaining summed columns mirror the sample's ``=SUM(<col>13:<col>N)``
        # formulas in formula mode, or the pre-evaluated literal otherwise.
        self._put_total(ws, row, COL_AMOUNT, "L", first, last, totals.total_amount_usd)
        self._put_total(
            ws, row, COL_RATE_PER_USD, "M", first, last,
            _sum_optional(ln.purchase_inr for ln in lines),
        )
        self._put_total(
            ws, row, COL_CUSTOM_ASS_VALUE, "N", first, last, totals.total_assessable_value
        )
        self._put_total(
            ws, row, COL_LAND_COST_WITHOUT_GST, "O", first, last,
            totals.total_land_cost_excl_gst,
        )
        self._put_total(
            ws, row, COL_TOTAL_CUSTOM_DUTY, "P", first, last, totals.total_customs_duty
        )
        self._put_total(ws, row, COL_GST, "Q", first, last, totals.total_igst)
        self._put_total(
            ws, row, COL_TOTAL_CUSTOM_DUTY_2, "S", first, last,
            _sum_optional(ln.combined_duty for ln in lines),
        )
        self._put_total(
            ws, row, COL_CUST_AIDC, "U", first, last,
            _sum_optional(ln.cust_aidc for ln in lines),
        )
        self._put_total(
            ws, row, COL_SURCHARGE, "W", first, last,
            _sum_optional(ln.sws_amount for ln in lines),
        )
        self._put_total(
            ws, row, COL_LAND_COST_WITH_GST, "X", first, last,
            totals.total_land_cost_incl_gst,
        )

    # ------------------------------------------------------------------
    # Auxiliary template sections (labels only; data cells empty)
    # ------------------------------------------------------------------
    def _write_aux_templates(self, ws: Worksheet, shift: int, totals_row: int) -> None:
        """Write the auxiliary section labels verbatim at their (shifted) positions.

        Reproduces character-for-character (Req 8.3): the ``total usd`` summary
        (B71 label + E71 value mirroring ``=L61``), the ``DETAILS AS PER
        CHALLANS`` and ``DETAILS AS PER TALLY`` section titles with their
        column-header rows, the C&F detail block row labels (B72-B86), and the
        ``CLEARANCE AND FORWARDING INVOICE`` title with its header row. When the
        line-item count overflows the template's data region these positions are
        shifted down by ``shift`` rows so they sit below the item rows rather
        than colliding with them; with no overflow (``shift == 0``) they remain
        at the sample's exact cell positions (Req 8.2). The data cells within
        these sections are intentionally left empty for Milestone 1.
        """
        for coordinate, label in AUX_LABELS.items():
            ws[_shift_coord(coordinate, shift)] = label

        # 'total usd' value cell: the sample stores '=L61'; mirror that with a
        # live formula in formula mode, or the evaluated literal otherwise. The
        # Totals_Row L cell holds the identical value/formula.
        target = _shift_coord(AUX_TOTAL_USD_CELL, shift)
        if self.use_formulas:
            ws[target] = f"=L{totals_row}"
        else:
            total_usd = ws.cell(row=totals_row, column=COL_AMOUNT).value
            if total_usd is not None:
                ws[target] = total_usd

    # ------------------------------------------------------------------
    # Small writing helpers
    # ------------------------------------------------------------------
    def _put_computed(
        self, ws: Worksheet, row: int, col: int, value: object | None, formula: str
    ) -> None:
        """Write a computed cell as a formula or literal, blank when ``None``.

        In formula mode the live Excel ``formula`` is written (mirroring the
        sample workbook); otherwise the pre-evaluated ``value`` literal is
        written. A ``None`` value (missing/non-numeric input, Req 6.13) leaves
        the cell blank in both modes, so a formula is never written for a cell
        whose inputs were absent (which would otherwise read blanks as 0).
        """
        if value is None:
            return
        if self.use_formulas:
            ws.cell(row=row, column=col, value=formula)
        else:
            ws.cell(row=row, column=col, value=value)

    def _put_total(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        col_letter: str,
        first: int,
        last: int,
        value: object | None,
    ) -> None:
        """Write a Totals_Row cell as ``=SUM(<col>first:<col>last)`` or a literal.

        In formula mode a ``SUM`` over the actual data rows is written (matching
        the sample's ``=SUM(L13:L59)`` style); with no data rows
        (``last < first``) or in literal mode the pre-evaluated ``value`` is
        written instead.
        """
        if self.use_formulas and last >= first:
            ws.cell(row=row, column=col, value=f"=SUM({col_letter}{first}:{col_letter}{last})")
        else:
            self._set_cell(ws, row, col, value)

    @staticmethod
    def _num(value: float) -> str:
        """Render a numeric literal for embedding in a formula (no exponent).

        Integers render without a trailing ``.0`` (e.g. ``95`` not ``95.0``) and
        finite floats render in plain decimal form so the formula text matches
        the sample's style (``=L13*95.3``).
        """
        if value is None:
            return "0"
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return repr(value)

    @staticmethod
    def _set_cell(ws: Worksheet, row: int, col: int, value: object | None) -> None:
        """Write ``value`` at ``(row, col)`` only when it is not ``None``.

        Leaving ``None`` unwritten keeps no-source/flagged cells truly empty (no
        characters, spaces, or placeholder text) per Req 4.8/4.9/5.11/8.6.
        """
        if value is not None:
            ws.cell(row=row, column=col, value=value)

    @staticmethod
    def _set_if_value(ws: Worksheet, coordinate: str, value: object | None) -> None:
        """Write ``value`` at ``coordinate`` only when it is not ``None``."""
        if value is not None:
            ws[coordinate] = value
