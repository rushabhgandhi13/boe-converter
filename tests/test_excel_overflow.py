"""Dynamic positioning: totals row + auxiliary sections shift below the items.

The bundled style template carries a styled data region of rows 13..60, a totals
row at 61, and the auxiliary sections (``DETAILS AS PER CHALLANS`` / ``DETAILS AS
PER TALLY`` at row 71, the C&F block, and ``CLEARANCE AND FORWARDING INVOICE`` at
row 94). When a Bill of Entry has more line items than the template's data
region, the item rows would previously overrun and collide with the fixed totals
row and the auxiliary tables.

These tests verify that for a document with many line items:

- the item rows remain dense from row 13 (Sr. no. 1..N) with no gaps;
- the totals row is written immediately below the last item row (not at the
  fixed row 61) and carries the invoice-amount total;
- every auxiliary label sits strictly below the totals row, so no label cell is
  overwritten by an item row and no item row is overwritten by a label;
- the auxiliary section merges follow the labels to their shifted positions.

It also asserts the no-overflow path (a small document) is unchanged: totals stay
at row 61 and the aux sections stay at their sample positions.
"""

from __future__ import annotations

from boe_converter.excel_writer import (
    AUX_LABELS,
    AUX_TOTAL_USD_CELL,
    COL_AMOUNT,
    COL_SR_NO,
    ITEM_TABLE_FIRST_DATA_ROW,
    TOTALS_ROW,
    ExcelGenerator,
    _shift_coord,
)
from boe_converter.models import (
    ComputedDocument,
    ComputedLine,
    HeaderBlock,
    LineItem,
    RawValue,
    ReviewFlagSet,
    Totals,
)
from openpyxl.utils.cell import coordinate_to_tuple


def _raw(value) -> RawValue:
    return RawValue(raw_text=str(value), parsed=value)


def _header() -> HeaderBlock:
    return HeaderBlock(
        company_name="Gemini Unicom LLP",
        party_name=_raw("ACME"),
        usd_rate=90.0,
        details=_raw("CO-32"),
        invoice_no=_raw("INV-1"),
        invoice_date=_raw("01/01/2026"),
        be_no=_raw("BE-1"),
        be_date=_raw("01/01/2026"),
        bl_no=_raw("BL-1"),
        bl_date=_raw("01/01/2026"),
        invoice_amount=_raw(100.0),
        invoice_currency=_raw("USD"),
        package_count=_raw(1357),
        container_details=_raw("CONT-1"),
    )


def _line(serial: int) -> ComputedLine:
    item = LineItem(
        item_serial=serial,
        cth_hsn=_raw("1234"),
        description=_raw(f"ITEM-{serial}"),
        unit_price_usd=_raw(2.0),
        quantity=_raw(3.0),
        unit=_raw("PCS"),
        assessable_value=_raw(10.0),
        bcd_rate=_raw(0.1),
        bcd_amount=_raw(1.0),
        igst_rate=_raw(0.18),
        total_duty=_raw(2.0),
    )
    return ComputedLine(source=item)


def _build(n_items: int):
    lines = [_line(s) for s in range(1, n_items + 1)]
    calc_lines = lines  # ComputedLine values are not needed for positioning here
    doc = ComputedDocument(
        header=_header(),
        lines=calc_lines,
        totals=Totals(total_amount_usd=float(n_items)),
    )
    wb = ExcelGenerator().build_workbook(doc, ReviewFlagSet())
    return wb["Sheet1"], doc


def test_overflow_shifts_totals_and_aux_below_items():
    """80 items: totals + aux move below the item block, no overlap."""
    n = 80
    ws, _ = _build(n)
    first = ITEM_TABLE_FIRST_DATA_ROW

    # Dense Sr. no. 1..N from row 13, with the row after the block empty.
    sr_values = [ws.cell(row=first + i, column=COL_SR_NO).value for i in range(n)]
    assert sr_values == list(range(1, n + 1))
    last_item_row = first + n - 1  # 92
    assert ws.cell(row=last_item_row + 1, column=COL_SR_NO).value is None

    # Totals row is immediately below the last item row (not the fixed row 61).
    shift = last_item_row - 60
    expected_totals_row = TOTALS_ROW + shift
    assert expected_totals_row == last_item_row + 1
    amount_total = ws.cell(row=expected_totals_row, column=COL_AMOUNT).value
    assert isinstance(amount_total, (int, float)) and amount_total > 0

    # Every auxiliary label sits strictly below the totals row (no collision
    # with any item row) and matches the verbatim template label at its
    # shifted coordinate.
    for coord, label in AUX_LABELS.items():
        shifted = _shift_coord(coord, shift)
        row, _col = coordinate_to_tuple(shifted)
        assert row > expected_totals_row, f"{coord}->{shifted} overlaps items/totals"
        assert ws[shifted].value == label

    # The 'total usd' value cell tracks the shifted totals amount.
    assert ws[_shift_coord(AUX_TOTAL_USD_CELL, shift)].value == amount_total

    # No item row (13..last_item_row) accidentally carries an aux label.
    aux_labels = set(AUX_LABELS.values())
    for r in range(first, last_item_row + 1):
        for c in range(1, 26):
            assert ws.cell(row=r, column=c).value not in aux_labels

    # The shifted aux section merges exist at their new positions.
    merged = {str(rng) for rng in ws.merged_cells.ranges}
    assert _shift_coord("G71", shift) + ":" + _shift_coord("O71", shift) in merged
    assert _shift_coord("Q71", shift) + ":" + _shift_coord("X71", shift) in merged
    assert _shift_coord("G94", shift) + ":" + _shift_coord("O94", shift) in merged


def test_no_overflow_keeps_sample_positions():
    """A small document keeps the totals at row 61 and aux at sample rows."""
    ws, _ = _build(10)

    # Totals at the fixed sample row, with the amount total present.
    amount_total = ws.cell(row=TOTALS_ROW, column=COL_AMOUNT).value
    assert isinstance(amount_total, (int, float)) and amount_total > 0

    # Aux labels remain at their original (unshifted) coordinates.
    for coord, label in AUX_LABELS.items():
        assert ws[coord].value == label

    # Original aux merges are intact.
    merged = {str(rng) for rng in ws.merged_cells.ranges}
    assert "G71:O71" in merged
    assert "Q71:X71" in merged
    assert "G94:O94" in merged


def test_boundary_exactly_fills_region():
    """48 items exactly fills rows 13..60; totals stay at row 61 (no shift)."""
    n = 48
    ws, _ = _build(n)
    first = ITEM_TABLE_FIRST_DATA_ROW
    assert ws.cell(row=first + n - 1, column=COL_SR_NO).value == n  # row 60
    assert ws.cell(row=TOTALS_ROW, column=COL_SR_NO).value is None  # row 61 clear of items
    assert ws.cell(row=TOTALS_ROW, column=COL_AMOUNT).value > 0
