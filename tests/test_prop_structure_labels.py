"""Property-based test for output structure and fixed labels (task 7.6).

Property 12: Output structure and fixed labels match the sample exactly.

**Validates: Requirements 5.12, 8.1, 8.2, 8.3**

For any document, the output workbook contains exactly one worksheet named
``Sheet1``; the Header_Block fixed labels occupy their cells in ``D1:G8``; the
Item_Table header is at row 12 with each column label a character-for-character
match of the sample; the Totals_Row sits at the sample's fixed position (row
61); and the auxiliary section labels appear at the sample's exact cell
positions. These structural facts are invariant across every input - they do
not depend on the line-item data - so the test generates varied
``ComputedDocument``s (different item counts, including 0 items) and asserts the
same structure holds for all of them.

The expected label maps (``ITEM_TABLE_HEADERS``, ``AUX_LABELS``) are imported
from ``boe_converter.excel_writer`` and asserted to be written verbatim into the
workbook at their mapped cells, so the test confirms the writer actually emits
them at the right positions (not merely that the constants exist).
"""

from __future__ import annotations

from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl.utils import get_column_letter

from boe_converter.calculator import ValueCalculator
from boe_converter.excel_writer import (
    AUX_LABELS,
    ITEM_TABLE_HEADER_ROW,
    ITEM_TABLE_HEADERS,
    TOTALS_ROW,
    ExcelGenerator,
)
from boe_converter.models import (
    ExtractedDocument,
    HeaderBlock,
    LineItem,
    RawValue,
    ReviewFlagSet,
)

# Header_Block fixed labels at their mapped cells (design.md cell map, D1:G8).
# These are written unconditionally by the writer regardless of input, so they
# are part of the fixed output structure (Req 8.3).
HEADER_BLOCK_LABELS: dict[str, str] = {
    "D1": "Company name",
    "D2": "Party Name",
    "F2": "USD Rate",
    "D3": "Details",
    "F3": "USD Amt",
    "D4": "Invoice No",
    "F4": "Inv Date",
    "D5": "BE No",
    "F5": "BE Date",
    "D6": "B/L NO",
    "F6": "B/L DATE",
    "D7": "Eway bill no",
    "F7": "Eway bill date",
    "D8": "RETTENCE DATE",
    "F8": "RETTENCE RATE",
}

# ---------------------------------------------------------------------------
# Generators
# ---------------------------------------------------------------------------

_MONEY = st.floats(min_value=-1e6, max_value=1e6, allow_nan=False, allow_infinity=False)
_RATE = st.floats(min_value=0.0, max_value=1.0, allow_nan=False, allow_infinity=False)
_USD_RATE = st.floats(min_value=0.0, max_value=200.0, allow_nan=False, allow_infinity=False)

# Text that round-trips verbatim through openpyxl (printable, no leading '=').
_TEXT = st.text(
    alphabet=st.characters(min_codepoint=32, max_codepoint=126, blacklist_characters="="),
    min_size=1,
    max_size=20,
)


def _num(value: float) -> RawValue:
    return RawValue(raw_text=repr(value), parsed=value)


def _str(value: str) -> RawValue:
    return RawValue(raw_text=value, parsed=value)


@st.composite
def _line_items(draw):
    """LineItems with unique, ascending serials; count includes 0 (Req 7.3)."""

    count = draw(st.integers(min_value=0, max_value=60))
    items: list[LineItem] = []
    for serial in range(1, count + 1):
        items.append(
            LineItem(
                item_serial=serial,
                cth_hsn=_str(draw(_TEXT)),
                description=_str(draw(_TEXT)),
                unit_price_usd=_num(draw(_MONEY)),
                quantity=_num(draw(_MONEY)),
                unit=_str(draw(st.sampled_from(("DOZ", "PCS", "KGS", "SET")))),
                assessable_value=_num(draw(_MONEY)),
                bcd_rate=_num(draw(_MONEY)),
                bcd_amount=_num(draw(_MONEY)),
                igst_rate=_num(draw(_RATE)),
                total_duty=_num(draw(_MONEY)),
            )
        )
    return items


@st.composite
def _documents(draw):
    """An ExtractedDocument with a populated header and a varied item list."""

    usd_rate = draw(_USD_RATE)
    header = HeaderBlock(
        company_name=draw(_TEXT),
        party_name=_str(draw(_TEXT)),
        usd_rate=usd_rate,
        details=_str(draw(_TEXT)),
        invoice_no=_str(draw(_TEXT)),
        invoice_date=_str(draw(_TEXT)),
        be_no=_str(draw(_TEXT)),
        be_date=_str(draw(_TEXT)),
        bl_no=_str(draw(_TEXT)),
        bl_date=_str(draw(_TEXT)),
        invoice_amount=_num(draw(_MONEY)),
        invoice_currency=_str("USD"),
        package_count=_num(draw(st.floats(
            min_value=0, max_value=1e6, allow_nan=False, allow_infinity=False
        ))),
        container_details=_str(draw(_TEXT)),
    )
    doc = ExtractedDocument(
        header=header,
        line_items=draw(_line_items()),
        declared_item_count=None,
        flags=[],
    )
    return doc, usd_rate


# ---------------------------------------------------------------------------
# Property 12
# ---------------------------------------------------------------------------


@given(payload=_documents())
@settings(deadline=None, max_examples=25)
def test_output_structure_and_fixed_labels_match_sample(payload):
    """Structure and fixed labels are invariant across all documents."""

    doc, usd_rate = payload
    computed = ValueCalculator().compute(doc, usd_rate)
    wb = ExcelGenerator().build_workbook(computed, ReviewFlagSet())

    # --- Req 8.1: exactly one worksheet, named "Sheet1" -----------------
    assert wb.sheetnames == ["Sheet1"]
    ws = wb["Sheet1"]

    # --- Req 8.3: Header_Block fixed labels at their D1:G8 cells ---------
    for coordinate, label in HEADER_BLOCK_LABELS.items():
        assert ws[coordinate].value == label

    # --- Req 5.12: Item_Table header at row 12, char-for-char labels ----
    for col, label in ITEM_TABLE_HEADERS.items():
        cell = ws.cell(row=ITEM_TABLE_HEADER_ROW, column=col)
        assert cell.value == label, (
            f"header {get_column_letter(col)}{ITEM_TABLE_HEADER_ROW} "
            f"expected {label!r} got {cell.value!r}"
        )

    # --- Req 8.2: Totals_Row sits at the sample's fixed position (61) ----
    # The totals row is always row 61 regardless of item count; at least one
    # summed column (L = total invoice amount USD) carries the total there.
    from boe_converter.excel_writer import COL_AMOUNT

    assert TOTALS_ROW == 61
    assert ws.cell(row=TOTALS_ROW, column=COL_AMOUNT).value == computed.totals.total_amount_usd

    # --- Req 8.3: auxiliary section labels at their exact cells ---------
    for coordinate, label in AUX_LABELS.items():
        assert ws[coordinate].value == label, (
            f"aux {coordinate} expected {label!r} got {ws[coordinate].value!r}"
        )
