"""Streamlit front-end for the Bill of Entry converter.

A single-view, three-step workflow (no tabs):

  Step 1  Bill of Entry PDF  ->  CTN Excel (download).
  Step 2  Map each line's "AS PER TALLY NAME" using a dropdown of names stored
          in Neon, then download the updated Excel.
  Step 3  Excel (from Step 1/2 or an uploaded file)  ->  Tally Purchase-voucher
          JSON. Buyer/seller identity is filled FROM THE BOE with an optional
          override from stored data.

Reference data (stock-item names, buyer records, seller records) lives in a
**Neon Postgres** database so it survives Streamlit Cloud restarts. If the
database is not reachable the app flags it and the mapping / JSON steps are
disabled - there is no local-file fallback.

Run locally:    streamlit run streamlit_app.py
Deploy:         push to GitHub and point Streamlit Community Cloud at this file.

NOTE: the Excel conversion itself is unchanged; only the Tally-name mapping and
JSON generation use the stored data.
"""

from __future__ import annotations

import hmac
import inspect
import io
import json
import os

import streamlit as st
from openpyxl import load_workbook

from boe_converter.excel_writer import (
    COL_AS_PER_TALLY_NAME,
    COL_DESCRIPTION,
    COL_HSN_CODE,
    COL_QTY,
    COL_SR_NO,
    COL_UNIT,
    ITEM_TABLE_FIRST_DATA_ROW,
)
from boe_converter.orchestrator import ConversionOrchestrator

st.set_page_config(page_title="Bill of Entry Converter", page_icon="📄", layout="wide")


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
def _check_password() -> bool:
    """Gate the app behind a shared password stored in Streamlit secrets."""
    if st.session_state.get("auth_ok"):
        return True

    expected = st.secrets.get("app_password")
    if not expected:
        st.error(
            "This app is not configured. Set `app_password` in the app's "
            "Streamlit secrets (Manage app → Settings → Secrets)."
        )
        st.stop()

    st.title("🔒 Bill of Entry Converter")
    with st.form("login", clear_on_submit=True):
        entered = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Sign in")

    if submitted:
        if hmac.compare_digest(str(entered), str(expected)):
            st.session_state["auth_ok"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False


if not _check_password():
    st.stop()


# ---------------------------------------------------------------------------
# Shared resources
# ---------------------------------------------------------------------------
@st.cache_resource
def _orchestrator() -> ConversionOrchestrator:
    """A single shared orchestrator (holds the in-memory download store)."""
    budget = float(os.environ.get("BOE_TIME_BUDGET_SECONDS", "600"))
    return ConversionOrchestrator(time_budget_seconds=budget)


def _database_url() -> str | None:
    """Resolve the Neon connection string from secrets/env (secret wins)."""
    try:
        val = st.secrets.get("DATABASE_URL")
    except Exception:
        val = None
    return val or os.environ.get("DATABASE_URL") or os.environ.get("NEON_DATABASE_URL")


@st.cache_resource
def _store():
    """Connect to Neon and ensure the schema exists.

    Returns the store on success, or ``(None, error_message)`` never - instead
    the caller uses :func:`_get_store` which surfaces a clear error.
    """
    from boe_converter.tally_store import TallyStore

    store = TallyStore(dsn=_database_url())
    store.init_schema()
    return store


def _get_store():
    """Return the connected store or ``None`` (after showing an error)."""
    from boe_converter.tally_store import StoreError

    try:
        return _store()
    except StoreError as exc:
        st.error(f"Stored-data database unavailable: {exc}")
        st.caption(
            "The mapping and JSON steps need the Neon database. Set `DATABASE_URL` "
            "in the app secrets to your Neon Postgres URL "
            "(postgresql://…?sslmode=require), then reboot the app."
        )
        return None


# ---------------------------------------------------------------------------
# Sidebar: manage stored data (Neon)
# ---------------------------------------------------------------------------
def _render_data_manager() -> None:
    from boe_converter.tally_store import BuyerRecord, SellerRecord, StoreError

    st.sidebar.header("📇 Stored data (Neon)")
    store = _get_store()
    if store is None:
        st.sidebar.warning("Database not connected.")
        return
    st.sidebar.success("Database connected.")

    # --- Stock item names ---
    with st.sidebar.expander("Stock item names", expanded=False):
        try:
            names = store.list_stock_items()
        except StoreError as exc:
            st.error(str(exc))
            names = []
        new_name = st.text_input("Add a stock name", key="add_stock")
        if st.button("Add name", key="btn_add_stock"):
            if new_name.strip() and store.add_stock_item(new_name):
                st.success(f"Added “{new_name.strip()}”.")
                st.rerun()
            else:
                st.info("Empty or already present.")
        if names:
            to_del = st.selectbox("Delete a name", ["—", *names], key="del_stock")
            if st.button("Delete name", key="btn_del_stock") and to_del != "—":
                store.delete_stock_item(to_del)
                st.rerun()
        st.caption(f"{len(names)} stored name(s).")

    # --- Buyers ---
    with st.sidebar.expander("Buyer records", expanded=False):
        try:
            buyers = store.list_buyers()
        except StoreError as exc:
            st.error(str(exc))
            buyers = []
        with st.form("add_buyer_form", clear_on_submit=True):
            b_name = st.text_input("Buyer name")
            b_gstin = st.text_input("GSTIN")
            b_state = st.text_input("State")
            b_pin = st.text_input("Pincode")
            b_addr = st.text_area("Address (one line per row)")
            if st.form_submit_button("Save buyer"):
                if b_name.strip():
                    store.add_buyer(
                        BuyerRecord(
                            name=b_name.strip(),
                            gstin=b_gstin.strip(),
                            state=b_state.strip(),
                            pincode=b_pin.strip(),
                            address_lines=[l for l in b_addr.splitlines() if l.strip()],
                        )
                    )
                    st.success("Buyer saved.")
                    st.rerun()
        if buyers:
            names_b = [b.name for b in buyers]
            to_del_b = st.selectbox("Delete a buyer", ["—", *names_b], key="del_buyer")
            if st.button("Delete buyer", key="btn_del_buyer") and to_del_b != "—":
                store.delete_buyer(to_del_b)
                st.rerun()
        st.caption(f"{len(buyers)} stored buyer(s).")

    # --- Sellers ---
    with st.sidebar.expander("Seller records", expanded=False):
        try:
            sellers = store.list_sellers()
        except StoreError as exc:
            st.error(str(exc))
            sellers = []
        with st.form("add_seller_form", clear_on_submit=True):
            s_name = st.text_input("Seller name")
            s_country = st.text_input("Country")
            s_addr = st.text_area("Address (one line per row)")
            if st.form_submit_button("Save seller"):
                if s_name.strip():
                    store.add_seller(
                        SellerRecord(
                            name=s_name.strip(),
                            country=s_country.strip(),
                            address_lines=[l for l in s_addr.splitlines() if l.strip()],
                        )
                    )
                    st.success("Seller saved.")
                    st.rerun()
        if sellers:
            names_s = [s.name for s in sellers]
            to_del_s = st.selectbox("Delete a seller", ["—", *names_s], key="del_seller")
            if st.button("Delete seller", key="btn_del_seller") and to_del_s != "—":
                store.delete_seller(to_del_s)
                st.rerun()
        st.caption(f"{len(sellers)} stored seller(s).")


# ---------------------------------------------------------------------------
# Helpers shared by the steps
# ---------------------------------------------------------------------------
def _num(rv) -> str:
    if rv is None or getattr(rv, "is_missing", False):
        return ""
    parsed = getattr(rv, "parsed", None)
    if isinstance(parsed, (int, float)):
        return f"{parsed:g}"
    return str(getattr(rv, "raw_text", "") or "")


def _line_rows(computed, tally_map: dict):
    """Rows for the Step-2 editor from a ComputedDocument."""
    rows = []
    for line in computed.lines:
        src = line.source
        rows.append(
            {
                "Sr": src.item_serial,
                "Description": _num(src.description) or (src.description.raw_text or ""),
                "HSN": _num(src.cth_hsn) or (src.cth_hsn.raw_text or ""),
                "Qty": _num(src.quantity),
                "Unit": _num(src.unit) or (src.unit.raw_text or ""),
                "As per Tally name": tally_map.get(src.item_serial),
            }
        )
    return rows


def _apply_name_overrides(computed, mapping: dict):
    """Return a ComputedDocument whose line descriptions use the mapped names."""
    from dataclasses import replace

    from boe_converter.models import RawValue

    new_lines = []
    for line in computed.lines:
        name = mapping.get(line.source.item_serial)
        if name and name.strip():
            new_src = replace(
                line.source, description=RawValue(raw_text=name.strip(), parsed=name.strip())
            )
            new_lines.append(replace(line, source=new_src))
        else:
            new_lines.append(line)
    return replace(computed, lines=new_lines)


def _write_tally_names(workbook_bytes: bytes, mapping: dict) -> bytes:
    """Write the chosen Tally names into column D of the CTN workbook."""
    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb.active
    row = ITEM_TABLE_FIRST_DATA_ROW
    blank = 0
    while row < 2000:
        sr = ws.cell(row=row, column=COL_SR_NO).value
        desc = ws.cell(row=row, column=COL_DESCRIPTION).value
        is_data = isinstance(sr, (int, float)) and not isinstance(sr, bool)
        if not is_data and not (isinstance(desc, str) and desc.strip()):
            blank += 1
            if blank >= 3:
                break
            row += 1
            continue
        blank = 0
        if is_data and int(sr) in mapping and mapping[int(sr)]:
            ws.cell(row=row, column=COL_AS_PER_TALLY_NAME).value = mapping[int(sr)]
        row += 1
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Page
# ---------------------------------------------------------------------------
st.title("Bill of Entry → CTN Excel → Tally JSON")
_render_data_manager()

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ===========================================================================
# STEP 1 — Bill of Entry PDF -> CTN Excel
# ===========================================================================
st.header("Step 1 · Bill of Entry → Excel")
st.caption("Upload an ICEGATE Bill of Entry PDF and download the CTN workbook.")

col_a, col_b = st.columns(2)
with col_a:
    uploaded = st.file_uploader("Bill of Entry PDF", type=["pdf"], key="boe_pdf")
    invoice_uploaded = st.file_uploader(
        "Invoice / Packing List PDF (optional)",
        type=["pdf"],
        key="inv_pdf",
        help=(
            "Optional. If provided, per-line carton counts (CTN, column G) are read "
            "from the invoice's TOTAL CTNS column and matched by serial number."
        ),
    )
with col_b:
    usd_rate = st.number_input(
        "USD rate", min_value=0.0, value=95.30, step=0.01, format="%.2f",
        help="The USD→INR conversion rate applied to every line item.",
    )

if st.button("Convert to Excel", type="primary", disabled=uploaded is None):
    raw = uploaded.getvalue()
    invoice_raw = invoice_uploaded.getvalue() if invoice_uploaded is not None else None
    orchestrator = _orchestrator()

    supports_invoice = (
        "invoice_raw" in inspect.signature(orchestrator.convert).parameters
    )
    with st.spinner("Converting…"):
        if supports_invoice:
            result = orchestrator.convert(
                raw, uploaded.name, float(usd_rate), invoice_raw=invoice_raw
            )
        else:
            result = orchestrator.convert(raw, uploaded.name, float(usd_rate))

    if not result.ok:
        st.error(f"{result.error_code}: {result.message}")
    else:
        st.session_state["last_token"] = result.download_token
        st.session_state["last_usd_rate"] = float(usd_rate)
        st.session_state.pop("tally_map", None)  # reset mapping for a new BOE
        summary = result.summary
        if result.output_complete:
            st.success(f"Converted {summary.line_items_extracted} line items.")
        else:
            st.warning(
                f"Converted with an item-count mismatch: extracted "
                f"{summary.line_items_extracted}, BOE declared "
                f"{summary.declared_item_count}. Review before use."
            )

# Persisted download for Step 1 (survives reruns within the session).
_token = st.session_state.get("last_token")
if _token:
    wb_bytes = _orchestrator().get_download(_token)
    if wb_bytes:
        st.download_button(
            "⬇ Download CTN workbook (.xlsx)",
            data=wb_bytes,
            file_name="bill_of_entry.xlsx",
            mime=XLSX_MIME,
        )

st.divider()

# ===========================================================================
# STEP 2 — Map "AS PER TALLY NAME" via stored dropdown
# ===========================================================================
st.header("Step 2 · Map Tally stock names")
st.caption(
    "For each line, pick the exact Tally stock name (from your stored list) in "
    "the “As per Tally name” column, then download the updated Excel. Add new "
    "names in the sidebar first if they’re missing."
)

_computed = _orchestrator().get_computed(_token) if _token else None
if _computed is None:
    st.info("Convert a Bill of Entry in Step 1 to map its line items here.")
else:
    store = _get_store()
    stock_options = []
    if store is not None:
        from boe_converter.tally_store import StoreError

        try:
            stock_options = store.list_stock_items()
        except StoreError as exc:
            st.error(str(exc))

    # Pre-select from the current mapping so choices survive edits.
    rows = _line_rows(_computed, st.session_state.get("tally_map", {}))

    edited = st.data_editor(
        rows,
        key="tally_editor",
        use_container_width=True,
        hide_index=True,
        disabled=["Sr", "Description", "HSN", "Qty", "Unit"],
        column_config={
            "As per Tally name": st.column_config.SelectboxColumn(
                "As per Tally name",
                help="Exact Tally stock name (managed in the sidebar).",
                options=stock_options,
                required=False,
            )
        },
    )

    # Rebuild the serial -> chosen-name mapping from the editor.
    mapping = {}
    for r in edited:
        name = r.get("As per Tally name")
        if name:
            try:
                mapping[int(r["Sr"])] = str(name)
            except (TypeError, ValueError):
                pass
    st.session_state["tally_map"] = mapping

    unmapped = sum(1 for r in edited if not r.get("As per Tally name"))
    if unmapped:
        st.caption(f"{unmapped} line(s) still without a Tally name.")

    if _token:
        wb_bytes = _orchestrator().get_download(_token)
        if wb_bytes:
            st.download_button(
                "⬇ Download Excel with Tally names",
                data=_write_tally_names(wb_bytes, mapping),
                file_name="bill_of_entry_tally.xlsx",
                mime=XLSX_MIME,
            )

st.divider()

# ===========================================================================
# STEP 3 — Excel -> Tally Purchase-voucher JSON
# ===========================================================================
st.header("Step 3 · Generate Tally JSON")
st.caption(
    "Builds a Purchase voucher (line items grouped by IGST rate). Buyer and "
    "seller details are taken from the Bill of Entry; override them from stored "
    "records if needed."
)

from boe_converter.excel_reader import ExcelReadError, read_workbook
from boe_converter.tally_exporter import CompanyProfile, SellerProfile, TallyExporter

source = st.radio(
    "Excel source",
    ["Use the last conversion (with my Step 2 mapping)", "Upload an edited Excel"],
    index=0,
)

manual_doc = None
if source == "Upload an edited Excel":
    up = st.file_uploader("Edited CTN workbook (.xlsx)", type=["xlsx"], key="json_xlsx")
    if up is not None:
        try:
            manual_doc = read_workbook(up.getvalue())
        except ExcelReadError as exc:
            st.error(str(exc))


def _resolve_computed_for_json():
    """(ComputedDocument, usd_rate) from the chosen source, or (None, 0)."""
    if source == "Upload an edited Excel":
        if manual_doc is None:
            st.info("Upload an edited CTN workbook to continue.")
            return None, 0.0
        return manual_doc, manual_doc.header.usd_rate or st.session_state.get(
            "last_usd_rate", 0.0
        )
    if _computed is None:
        st.info("Convert a Bill of Entry in Step 1 (or upload an Excel above).")
        return None, 0.0
    mapping = st.session_state.get("tally_map", {})
    doc = _apply_name_overrides(_computed, mapping) if mapping else _computed
    return doc, st.session_state.get("last_usd_rate", _computed.header.usd_rate)


computed_for_json, rate = _resolve_computed_for_json()

# --- Buyer / seller override (defaults come FROM the BOE) ---
store = _get_store()
buyer_override = CompanyProfile()
seller_override = SellerProfile()
if computed_for_json is not None:
    hdr = computed_for_json.header

    def _rv_text(rv):
        if rv is None or getattr(rv, "is_missing", False):
            return ""
        p = getattr(rv, "parsed", None)
        return str(p) if isinstance(p, str) else (getattr(rv, "raw_text", "") or "")

    with st.expander("Buyer (importer) — from BOE, override if needed", expanded=False):
        buyer_names = []
        if store is not None:
            from boe_converter.tally_store import StoreError

            try:
                buyer_names = [b.name for b in store.list_buyers()]
            except StoreError as exc:
                st.error(str(exc))
        pick = st.selectbox("Use a stored buyer", ["(from BOE)", *buyer_names], key="pick_buyer")
        if pick != "(from BOE)" and store is not None:
            rec = store.find_buyer(pick)
            if rec:
                buyer_override = CompanyProfile(
                    name=rec.name or None,
                    gstin=rec.gstin or None,
                    state=rec.state or None,
                    pincode=rec.pincode or None,
                    address_lines=tuple(rec.address_lines) or None,
                )
        st.write(
            {
                "Name": buyer_override.name or hdr.company_name,
                "GSTIN": buyer_override.gstin or _rv_text(hdr.buyer_gstin),
                "State": buyer_override.state or _rv_text(hdr.buyer_state),
                "Pincode": buyer_override.pincode or _rv_text(hdr.buyer_pincode),
                "Address": list(buyer_override.address_lines or [])
                or _rv_text(hdr.buyer_address),
            }
        )

    with st.expander("Seller (supplier) — from BOE, override if needed", expanded=False):
        seller_names = []
        if store is not None:
            from boe_converter.tally_store import StoreError

            try:
                seller_names = [s.name for s in store.list_sellers()]
            except StoreError as exc:
                st.error(str(exc))
        pick_s = st.selectbox("Use a stored seller", ["(from BOE)", *seller_names], key="pick_seller")
        if pick_s != "(from BOE)" and store is not None:
            rec = store.find_seller(pick_s)
            if rec:
                seller_override = SellerProfile(
                    name=rec.name or None,
                    country=rec.country or None,
                    address_lines=tuple(rec.address_lines) or None,
                )
        st.write(
            {
                "Name": seller_override.name or _rv_text(hdr.party_name),
                "Country": seller_override.country or _rv_text(hdr.seller_country),
                "Address": list(seller_override.address_lines or [])
                or _rv_text(hdr.seller_address),
            }
        )

if st.button("Generate Tally JSON", type="primary", disabled=computed_for_json is None):
    exporter = TallyExporter(company=buyer_override, seller=seller_override)
    document = exporter.build(computed_for_json, float(rate))
    payload = json.dumps(document, ensure_ascii=True, indent=1).encode("utf-8")
    st.success("Tally Purchase-voucher JSON generated.")
    st.download_button(
        "⬇ Download Tally voucher JSON",
        data=payload,
        file_name="tally_purchase_voucher.json",
        mime="application/json",
    )
    with st.expander("Ledgers used in this voucher"):
        st.table([{"Ledger": n} for n in exporter.required_ledger_names(computed_for_json)])
